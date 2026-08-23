import re, json
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

SRC = '/opt/data/allnup-clone/public/pages/counsel.html'
OUT = '/opt/data/allnup-clone/public/assets/fee_table.xlsx'

html = open(SRC, encoding='utf-8').read()
m = re.search(r'<script id="worksheet-data" type="application/json">(.*?)</script>', html, re.S)
data = json.loads(m.group(1))
header = data[0]

cols = ['상품명', '모델명', '브랜드', '제품군', '규정']
fee_cols = [h for h in header if '수수료' in h]
idx = {h: header.index(h) for h in cols + fee_cols}

wb = Workbook()
ws = wb.active
ws.title = "fee"
out_header = cols + fee_cols
ws.append(out_header)

hf = Font(bold=True, color="FFFFFF")
hfill = PatternFill("solid", fgColor="2F66FF")
for c in range(1, len(out_header) + 1):
    cell = ws.cell(row=1, column=c)
    cell.font = hf
    cell.fill = hfill
    cell.alignment = Alignment(horizontal="center", vertical="center")

cnt = 0
for row in data[1:]:
    fees = [row[idx[h]] for h in fee_cols]
    if any(str(f).strip() for f in fees):
        out = [row[idx[h]] for h in cols] + fees
        ws.append(out)
        cnt += 1

widths = [28, 16, 8, 14, 14] + [12] * len(fee_cols)
for i, w in enumerate(widths, 1):
    ws.column_dimensions[chr(64 + i) if i <= 26 else 'A'].width = w

thin = Side(style="thin", color="DDDDDD")
border = Border(left=thin, right=thin, top=thin, bottom=thin)
for r in range(2, cnt + 2):
    for c in range(1, len(out_header) + 1):
        ws.cell(row=r, column=c).border = border
    for ci, h in enumerate(fee_cols):
        cell = ws.cell(row=r, column=len(cols) + ci + 1)
        val = str(cell.value).replace(',', '').strip()
        if val.isdigit():
            cell.number_format = '#,##0'
            cell.alignment = Alignment(horizontal="right")

ws.freeze_panes = "A2"
wb.save(OUT)
print(f"OK rows={cnt} fees={fee_cols}")
